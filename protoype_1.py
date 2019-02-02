# A porosity calculator designed to work with essentially any CT scan, assuming that the sample is relatively closed.
# Works using a shrinkwrap ROI approach to delineate the boundary of our sample within the image. Given a boundary,
# computing porosity simply involves taking a ratio of dark/light pixels. Does not produce accurate results if: images
# are in any color other than black and white, have lots of noisy pixels, sample has large gaps in its boundary.

import glob
import re
import cv2
import numpy as np
import time
import matplotlib.pyplot as plt
import random
import sys
import openpyxl
from openpyxl import load_workbook

file_location = None
loop_inc = 1  # Represents the step for the main loop, skips loop_inc - 1 images (Increase to speed up runtime)
width_inc = 1  # Represents the inc/decrement step in shape outliner (Increase to speed up runtime)
save_count = 0  # Keeps track of how many comparison images are saved (for naming purposes)


# Main control for the program. Reads in images, crops them to a smaller size, performs a ROI shrinkwrap, measures
# the porosity and records it to a excel file
def main():
	images = file_reader()  # Read in images

	wb_ws_save = excel_handler()  # Setup excel sheet
	wb = wb_ws_save[0]  # The current workbook
	ws = wb_ws_save[1]  # The current sheet within the workbook
	save_as = wb_ws_save[2]  # The title of the current workbook

	parameters = crop_parameters(images)    # Get the universal parameters for future cropping operations

	porosities = []

	start = time.time()

	# Iterate through the images and count the porosity
	for i in range(0, len(images), loop_inc):
		progress_tracker(i + 1, len(images))  # Shows user the programs progress
		total_img_pixels = np.count_nonzero(images[i])
		if total_img_pixels > 10:  # Exclude images that contain little to no white pixels
			crop_img = crop(images[i], parameters)  # Crop the images
			shape = shape_outliner(crop_img)
			total_shape_pixels = np.count_nonzero(shape)
			if total_shape_pixels != 0:  # Because some all black images are "sneaking through", do a second check!
				porosities.append(por_calc(total_img_pixels, total_shape_pixels))

	if len(porosities) != 0:
		porosity = np.average(np.array(porosities))
	else:
		porosity = 0

	print("Porosity: " + str(porosity))

	end = time.time()
	print("Time: " + str(end - start))

	data_writer(ws, porosity)  # Write the porosity to the excel file

	wb.save(save_as)  # Saves the excel file


# Reads in image files specified by the users
def file_reader():
	global file_location

	while True:
		file_location = input("Please specify the file path to where your images are stored: ")
		file_type = input("Please specify the file type. I.e bmp: ")
		files = (glob.glob(file_location + "/*." + file_type))  # Read all files in a folder using glob
		images = []
		for img_file in files:
			images.append(cv2.imread(img_file, 0))  # Turn the files into an array of images for easier access
		if len(images) == 0:
			print("The folder location or file type you selected were invalid, please try again")
			continue
		break

	return images


# Loads in a pre existing excel file or creates a new one for writing based on user input
def excel_handler():
	valid_file = False

	while not valid_file:
		old_excel = input("Would you like to use an existing excel file Y OR N: ")
		if old_excel == "Y" or old_excel == "y":  # User has selected a pre-existing workbook
			excel_file = file_location + "/" + input("Please specify the excel file name: ") + ".xlsx"
			try:
				wb = load_workbook(excel_file)
				ws = wb.active
				save_as = excel_file
				valid_file = True
			except IOError:
				print("Your specified file does not exist, please try again")

		elif old_excel == "N" or old_excel == "n":  # User has selected to create a new workbook
			wb = openpyxl.Workbook()
			save_as = input("What would you like to save the excel book as: ")
			if re.match("^[A-Za-z0-9_+@#^&()_,.!-]*$", save_as):
				save_as = file_location + "/" + save_as + ".xlsx"
				ws = wb.active
				valid_file = True
			else:
				print("This is not a valid file name, it contains a :,*,?,[,],\,/... please try again ")

		else:
			print("Sorry, I didn't understand that.")
			continue

	return wb, ws, save_as


# Given a excel sheet, data_writer writes the porosity previously computed to the excel sheet with a clean format
def data_writer(ws, porosity):
	i = 1  # We loop to find empty cells before outputting the results
	while ws.cell(row=i, column=1).value is not None and ws.cell(row=i + 1, column=1).value is not None:
		i += 3
	ws.cell(row=i, column=1).value = "Trial %d" % (i // 3)
	ws.cell(row=i + 1, column=1).value = "Porosity:"
	ws.cell(row=i + 1, column=2).value = porosity


# Randomly saves an image based on a tunable probability "odds" within functions
def random_saver():
	odds = 100  # Think of this as probability to save = 1/odds
	x = random.randrange(odds)
	return x == odds - 1


# Saves a plot showing a image and its corresponding shape outline, allows user to ensure proper operation of code
def image_saver(image1, image2, title1, title2):
	global save_count

	plt.clf()
	plt.subplot(2, 1, 1)
	plt.title(title1)
	plt.imshow(image1, cmap="gray")
	plt.subplots_adjust(hspace=.5)  # Adjusts the spacing so the plots do not overlap
	plt.subplot(2, 1, 2)
	plt.title(title2)
	plt.imshow(image2, cmap="gray")
	plt.savefig(file_location + "\Comparison_Number_" + str(save_count) + ".png")
	save_count += 1


# Displays the progress of the program (percent wise) to the terminal. Inspired by:
# https://stackoverflow.com/questions/43515165/pycharm-python-console-printing-on-the-same-line-not-working-as-intended
def progress_tracker(completion, total):
	sys.stdout.write("\r{0}".format(str("%.2f" % (completion / total * 100)) + "% Complete"))
	sys.stdout.flush()
	if completion == total:
		print("\nDONE!")


# Given a set of images, crop parameters estimates an appropriate crop boundary. To do this, we first consider x images
# with the highest bright pixel count. We assume there is a correlation between # of bright pixels and the area of the
# sample in the image. Given these top images, we then test how far we can crop each image in each of the compass
# directions. We then take the smallest values from each direction and then apply a scaling factor to these values.
# By doing this computation, we are careful to not cut into our data when doing our cropping.
def crop_parameters(images):
	height = images[0].shape[0]
	width = images[0].shape[1]

	images_considered = 10
	if len(images) < images_considered:  # Ensures we do not try and consider more images then exist
		images_considered = len(images)

	top_images = np.zeros([images_considered, 2], np.int32)
	parameters = np.zeros([images_considered, 4], np.int32)
	final_parameters = np.zeros([1, 4], np.int32)

	# Approximates the images with the largest respective area (assume most white pixels = most area)
	for i in range(0, len(images)):
		if np.count_nonzero(images[i]) > min(top_images[:, 1]):
			top_images[np.argmin(top_images[:, 1])] = [i, np.count_nonzero(images[i])]

	# Generate the number of pixels required to reach the sample in all 4 directions which guides where to crop
	for i in range(0, images_considered):
		parameters[i, 0] = top_down_shutter_close(images[top_images[i, 0]], 0, width, 1)  # Top Down
		parameters[i, 1] = top_down_shutter_close(images[top_images[i, 0]], height - 1, width, -1)  # Bottom Up
		parameters[i, 2] = left_right_shutter_close(images[top_images[i, 0]], 0, height, 1)  # Left to Right
		parameters[i, 3] = left_right_shutter_close(images[top_images[i, 0]], width - 1, height, -1)  # Right to Left

	# Take the smallest value for each of the 4 directions (minimum distance before hitting data)
	for i in range(0, 4):
		final_parameters[0, i] = min(parameters[:, i])

	return final_parameters[0]


# Used to probe how far a 2D plane can go in the image before reaching the dataset, works from the top and bottom edges
def top_down_shutter_close(image, height_index, width, increment):
	counter = 0

	# The loop ensure we do not stop at a faulty value caused by image noise, must be more than 10 non zero pixels in plane
	while np.count_nonzero(image[height_index, 0:width]) < 10:
		height_index += increment
		counter += 1

	return counter


# Used to probe how far a 2D plane can go in the image before reaching the dataset, works from the left and right edges
def left_right_shutter_close(image, width_index, height, increment):
	counter = 0

	# The loop ensure we do not stop at a faulty value caused by image noise, must be more than 10 non zero pixels in plane
	while np.count_nonzero(image[0:height, width_index]) < 10:
		width_index += increment
		counter += 1

	return counter


# Given an image and a list of parameters, crops the image to a smaller dimension for faster future processing
def crop(img, dimensions):
	scale = 1.2
	crop_img = img[int(dimensions[0] // scale):img.shape[0] - int(dimensions[1] // scale),
			int(dimensions[2] // scale):img.shape[1] - int(dimensions[3] // scale)]

	# Randomly choose to save a comparison image (Ensure cropper works properly)
	# if random_saver():
	# 	image_saver(img, crop_img, "Original Image", "Cropped image")

	return crop_img


# Mimics a ROI shrink wrap procedure. Traces across original image until it encounters a white pixel, for each pixel
# we encounter this way, we write the corresponding pixel in the shape array black. In doing this, we outline the
# sample within each image, which can then be used to calculate porosity.
def shape_outliner(image):
	height = image.shape[0]
	width = image.shape[1]
	shape = np.ones((height, width), dtype=int)
	cur_i = image

	# Move right to left across image moving down
	for j in range(0, height):
		width_index = width - 1  # Right hand side of image
		height_index = j  # Top of image
		if np.count_nonzero(cur_i[height_index, 0:width]) == 0:  # Check if there are any white pixels in current row
			shape[height_index, 0:width] = 0
		else:
			while cur_i[
				height_index, width_index] == 0 and width_index > width_inc:  # Loop until white pixel or boundary found
				width_index -= width_inc
			shape[height_index, width_index + 1:width] = 0  # Color the pixels we just iterated through black in shape array

	# Move left to right moving down
	for j in range(0, height):
		width_index = 0  # Left hand side of image
		height_index = j  # Top of image
		if np.count_nonzero(cur_i[height_index, 0:width]) == 0:  # Check if there are any white pixels in current row
			shape[height_index, 0:width] = 0
		else:
			while cur_i[
				height_index, width_index] == 0 and width_index < width - width_inc:  # Loop until white pixel or boundary found
				width_index += width_inc
			shape[height_index, 0:width_index] = 0  # Color the pixels we just iterated through black in shape array

	# shape = cv2.medianBlur(shape.astype(np.float32), 5)  # Sooth images to remove border irregularities

	# Randomly choose to save a comparison image (Ensure shape outliner works properly)
	if random_saver():
		image_saver(image, shape, "Original Image", "Shape outlined image")

	return shape


# Returns the porosity given a count of bright pixels vs amount of total pixels
def por_calc(bright_pixels, total_pixels):
	return (1 - (bright_pixels / float(total_pixels))) * 100


main()
