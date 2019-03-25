import os
import cv2
import glob
import random
import sys
import numpy as np
import openpyxl
import threading
from tkinter import *
from tkinter.ttk import Progressbar
from openpyxl import load_workbook
from skimage.morphology import remove_small_objects
from multiprocessing.dummy import Pool as ThreadPool
from scipy.ndimage import label

################################Defining Global Variables (For GUI interfacing)#########################################

# Collections of three images to display to the user the effect of various operations
test_image = []
threshold_test_images = []
despeckle_test_images = []
cropped_test_images = []

# Images read in to the program, and the path that they came from
images = []
file_path = None

# Used to measure the volume of a sample, optional parameter
voxel_size = None

# Determine if images should be saved and if so, what their names should be
save_images = False
output_image_filename = None

# Variables used to create/save the excel file where values will be output
workbook = None
worksheet = None
excel_file_name = None

# Determine the users threshold and despeckle type
threshold_type = None
despeckle_type = None  # Despeckle type in a two tuple, first value is the choice, second value is the area

# Custom parameter for global means thresholding
global_means_thresh_value = 127

# Variables to determine if cropping should be performed, and if so, gets the parameters and scale that will be used
perform_crop = None
cparams = None
scale = None

# Used to exit the final loading bar (essentially killing tkinter root.mainloop())
still_loading = True
current_operation = None

width_inc = 1  # Represents the inc/decrement step in shape outliner (Increase to speed up runtime)


###################################Functions that interface with front end GUI##########################################

# Function called from front end to read in image files. Ensures a valid file format is selected and if so, makes use
# of multi-threading to speed up the process so the user does not have to wait overly long to continue in the frontend.
def file_reader(file_location):
	global test_image, images, file_path

	if file_location is not None and len(file_location) > 1:

		# Need to get file location into proper format for glob
		file_type = file_location.split(".")[-1]
		for i in range(len(file_location) - 1, -1, -1):
			if file_location[i] == "/":
				file_path = file_location[0:i]
				break

		file_location = file_path + "/*." + file_type
		glob_file = (glob.glob(file_location))  # Read all files in a folder using glob

		try:
			cv2.resize(cv2.imread(glob_file[0], 0), (1, 1))  # Tests the files read are images using cv2.resize
		except cv2.error:
			return False

		def multi_thread_file_read(file):
			return cv2.imread(file, 0)

		# Multi Thread image read to speed up the process
		pool = ThreadPool(4)
		images = pool.map(multi_thread_file_read, glob_file)
		pool.close()
		pool.join()

		test_image = []  # Reset global value so I do not append on top of old values

		# Get a subset of images for test images
		for i in range(0, len(images), len(images) // 3):
			test_image.append(images[i])

		return True

	return False


# Read in a pre-existing excel file TODO CREATE NEW WORKSHEET IF CURRENT WORKSHEET FORMAT IS OFF
def old_excel_reader(excel_file):
	global workbook, worksheet, excel_file_name

	if excel_file is not None and len(excel_file) > 1:  # Ensures the user has actually selected an excel file
		workbook = load_workbook(excel_file)
		worksheet = workbook.active
		excel_file_name = excel_file

		return True

	return False


# Creates a new excel file (Note, input validation is performed in the frontend as part of filedialog.asksaveasfilename
def new_excel_reader(excel_file):
	global workbook, worksheet, excel_file_name

	if excel_file is not None and len(excel_file) > 1:
		excel_file = excel_file + ".xlsx"
		workbook = openpyxl.Workbook()
		worksheet = workbook.active
		excel_file_name = excel_file

		return True

	return False


# Parses the users chosen image file name output into a more usable format (Allows me to append numbers to the filename)
def save_images_as(img_name):
	global output_image_filename

	if img_name is not None and len(img_name) > 1:
		for i in range(len(img_name) - 1, -1, -1):
			if img_name[i] == "/":
				output_image_path = img_name[i:len(img_name) + 1]
				output_image_filename = output_image_path.split(".")  # First Part is filename, second part is file type
				break
		return True

	return False

# NOTE: THRESHOLD CHOICE SELECTION IS DONE DIRECTLY IN THE FRONT END BY ASSIGNING VALUES TO GLOBAL VARIABLE, THIS IS
# DONE BECAUSE THERE IS MORE VARIABILITY AMONG THE DIFFERENT OPTIONS, EASIER THAN PASSING LOTS OF DIFFERENT VALUES


# Set the type of despeckeling to perform as well as the area to despeckle
def despeckle_choice(user_choice, area):
	global despeckle_type

	despeckle_type = (user_choice, int(round(float(area))))


# Toggle if cropping will be performed and if so, generate the cropping parameters (see crop_parameters for details)
def crop_choice(user_choice):
	global perform_crop, cparams

	if user_choice:
		perform_crop = True
		cparams = crop_parameters()
	else:
		perform_crop = False


# Generates the images the user will see when doing a image comparison for thresholding in the front end
def threshold_test_image_generator(user_choice):
	global threshold_test_images

	threshold_test_images = []  # Reset global array (Otherwise we just append over old images)

	for i in range(len(test_image)):
		if user_choice == 1:
			threshold_test_images.append(otsu_threshold(test_image[i]))
		elif user_choice == 2:
			threshold_test_images.append(global_threshold(test_image[i]))
		elif user_choice == 3:
			threshold_test_images.append(phansalkar_threshold(test_image[i]))


# Generates the images the user will see when doing a image comparison for despeckeling in the front end
def despeckle_test_image_generator(user_choice, area):
	global despeckle_test_images

	despeckle_test_images = []  # Reset global array (Otherwise we just append over old images)

	for i in range(len(test_image)):
		if user_choice == 1:
			despeckle_test_images.append(less_than_despeckle(test_image[i], area))
		elif user_choice == 2:
			despeckle_test_images.append(greater_than_despeckle(test_image[i], area))
		elif user_choice == 3:
			despeckle_test_images.append(less_than_despeckle(test_image[i], area))


# Generates the images the user will see when doing a image comparison for cropping in the front end
def crop_test_image_generator(user_choice):
	global cropped_test_images, scale

	cropped_test_images = []  # Reset global array
	scale = user_choice

	for i in range(len(test_image)):
		cropped_test_images.append(crop(despeckle_test_images[i], user_choice))


###################################BACK END ONLY BELOW#################################################################

# Main_flow acts as a main in that it delegates all tasks of the backend. First it processes the images using
# image_processor, after which said images are fed into the analyzer which returns the cumulative porosity and volume
# of the total dataset. These values are given to the data writer to save to an excel file. Finally, the images we
# processed can be saved to an output folder at which point the program is finished!
def main_flow():
	global current_operation

	processed_images = image_processor()

	results = analyze(processed_images)  # Results computes porosity/ surface area for each slice
	porosities = results[0]
	surface_area = results[1]

	if voxel_size is not None:
		volume = count_volume(surface_area)
	else:
		volume = "N/A"

	porosity = np.average(np.array(porosities))

	data_writer(worksheet, porosity, volume)
	workbook.save(excel_file_name)

	if save_images:
		current_operation = "Saving Images"
		output_folder = output_images()
		for i in range(len(images)):
			cv2.imwrite(output_folder + output_image_filename[0] + str(i) + "." + output_image_filename[1], images[i])

	print(excel_file_name)

	return


# Image processors handles all image processing procedures which includes thresholding, despeckeling and cropping.
# The despeckle and crop functionality are essentially complete, but may need to be reworked with UI implementation
def image_processor():
	global current_operation

	# Threshold all images
	current_operation = "Thresholding"
	for i in range(0, len(images)):
		if threshold_type == 1:
			images[i] = otsu_threshold(images[i])
		elif threshold_type == 2:
			images[i] = global_threshold(images[i])
		else:
			images[i] = phansalkar_threshold(images[i])

	# Despeckle all images
	current_operation = "Despeckeling"
	for i in range(0, len(images)):
		if despeckle_type[0] == 1 or 3:
			images[i] = less_than_despeckle(images[i], despeckle_type[1])
		else:
			images[i] = greater_than_despeckle(images[i], despeckle_type[1])

	# Crop all images
	current_operation = "Cropping"
	if perform_crop:
		for i in range(0, len(images)):
			images[i] = crop(images[i])

	return images


# Analyze preforms porosity and surface area measurements on binary images. First each image gets a ROI shrinkwrap
# which accurately delineates the sample boundary. The total pixels of this shrinkwrap are added a surface area holder
# which is later multiplied by voxel size to estimate object volume. Likewise a ratio between bright pixels in the
# shape outlined image vs original image is used to estimate porosity. These values are then returned to main.
def analyze(processed_images):
	global current_operation

	porosities = []
	surface_area = 0
	pimages = processed_images

	# Iterate through the images counting porosity and surface area for each slice
	current_operation = "Calculating Porosity"
	for i in range(0, len(pimages)):
		progress_tracker(i + 1, len(pimages), 'Porosity estimation ')  # Shows user the programs progress
		total_img_pixels = np.count_nonzero(pimages[i])
		if total_img_pixels > 10:  # Exclude images that contain little to no white pixels
			shape = shape_outliner(pimages[i])
			total_shape_pixels = np.count_nonzero(shape)
			porosities.append(por_calc(total_img_pixels, total_shape_pixels))
			surface_area += total_shape_pixels

	return porosities, surface_area


def otsu_threshold(image):
	_, thres_img = cv2.threshold(image, 127, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
	return thres_img


def global_threshold(image):
	_, thres_img = cv2.threshold(image, global_means_thresh_value, 255, cv2.THRESH_BINARY)
	return thres_img


# TODO IMPLEMENT PHANSALKER
def phansalkar_threshold(image):
	_, thres_img = cv2.threshold(image, 127, 255, cv2.THRESH_BINARY)
	return thres_img


# Despeckle images using a remove small objects function, size of objects is either user determined or found
# automatically via random experimentation in auto despeckle parameters
def less_than_despeckle(image, min_area):

	min_area = int(round(float(min_area)))
	grey_scale_conv_array = np.full((image.shape[0], image.shape[1]), 255)

	bool_arr = np.array(image, bool)
	despeckeled_img = remove_small_objects(bool_arr, min_area)

	grey_scale = np.array([a*b for a, b in zip(despeckeled_img, grey_scale_conv_array)], dtype=np.uint8)

	return grey_scale


# Despeckle images using a remove small objects function, size of objects is user determined, use an XOR function on two
# overlapping binary arrays (original image and less than despeckeled version) which simulates a greater than despeckle effect.
def greater_than_despeckle(image, min_area):

	min_area = int(round(float(min_area)))
	grey_scale_conv_array = np.full((image.shape[0], image.shape[1]), 255)

	bool_arr = np.array(image, bool)
	despeckeled_img = remove_small_objects(bool_arr, min_area)

	final_image = np.array(np.logical_xor(image, despeckeled_img), dtype=np.uint8)
	grey_scale = np.array([a*b for a, b in zip(final_image, grey_scale_conv_array)], dtype=np.uint8)

	return grey_scale


# Auto despeckle is designed to work with single object, low porosity samples (i.e salt scans). It works by
# iteratively increasing the min object despeckeled until only one object remains. We repeat this process on random
# images and take the average value after x trials.
def auto_despeckle_parameters():
	images_considered = 10
	min_area_array = np.zeros([images_considered])

	for i in range(0, images_considered):
		random_img = images[random.randrange(len(images))]  # Select a random image from the stack
		bool_arr = np.array(random_img, bool)  # Convert it into a boolean array (needed for small objects method)

		min_obj_size = 0
		num_features = -1

		# Loop increasing min object size until only one object is left in the image
		while num_features != 1 and min_obj_size < 200:
			min_obj_size += 10
			filtered_img = remove_small_objects(bool_arr, min_obj_size)
			null, num_features = label(filtered_img)

		min_area_array[i] = min_obj_size

	return int(np.average(min_area_array))


# Given a set of images, crop parameters estimates an appropriate crop boundary. To do this, we first consider x images
# with the highest bright pixel count. We assume there is a correlation between # of bright pixels and the area of the
# sample in the image. Given these top images, we then test how far we can crop each image in each of the compass
# directions. We then take the smallest values from each direction and then apply a scaling factor to these values.
# By doing this computation, we are careful to not cut into our data when doing our cropping.
def crop_parameters():
	height = images[0].shape[0]
	width = images[0].shape[1]

	images_considered = 10
	if len(images) < images_considered:  # Ensures we do not try and consider more images then exist
		images_considered = len(images)

	top_images = np.zeros([images_considered, 2], np.int32)
	parameters = np.zeros([images_considered, 4], np.int32)
	final_parameters = np.zeros([1, 4], np.int32)

	# Approximates the images with the largest respective area (assume most white pixels = most area)
	for i in range(0, len(images)):
		if np.count_nonzero(images[i]) > min(top_images[:, 1]):
			top_images[np.argmin(top_images[:, 1])] = [i, np.count_nonzero(images[i])]

	# Generate the number of pixels required to reach the sample in all 4 directions which guides where to crop
	for i in range(0, images_considered):
		parameters[i, 0] = top_down_shutter_close(images[top_images[i, 0]], 0, width, 1)  # Top Down
		parameters[i, 1] = top_down_shutter_close(images[top_images[i, 0]], height - 1, width, -1)  # Bottom Up
		parameters[i, 2] = left_right_shutter_close(images[top_images[i, 0]], 0, height, 1)  # Left to Right
		parameters[i, 3] = left_right_shutter_close(images[top_images[i, 0]], width - 1, height, -1)  # Right to Left

	# Take the smallest value for each of the 4 directions (minimum distance before hitting data)
	for i in range(0, 4):
		final_parameters[0, i] = min(parameters[:, i])

	return final_parameters[0]


# Used to probe how far a 2D plane can go in the image before reaching the dataset, works from the top and bottom edges
def top_down_shutter_close(image, height_index, width, increment):
	counter = 0

	# The loop ensure we do not stop at a faulty value caused by image noise, must be more than 10 non zero pixels in plane
	while np.count_nonzero(image[height_index, 0:width]) < 10 and height_index < image.shape[0]:
		height_index += increment
		counter += 1

	return counter


# Used to probe how far a 2D plane can go in the image before reaching the dataset, works from the left and right edges
def left_right_shutter_close(image, width_index, height, increment):
	counter = 0

	# The loop ensure we do not stop at a faulty value caused by image noise, must be more than 10 non zero pixels in plane
	while np.count_nonzero(image[0:height, width_index]) < 10 and width_index < image.shape[1]:
		width_index += increment
		counter += 1

	return counter


# Given an image and a list of parameters, crops the image to a smaller dimension for faster future processing
def crop(img, scale=1):

	dimensions = cparams

	crop_img = img[int(dimensions[0] // scale):img.shape[0] - int(dimensions[1] // scale),
	int(dimensions[2] // scale):img.shape[1] - int(dimensions[3] // scale)]

	return crop_img


# Mimics a ROI shrink wrap procedure. Traces across original image until it encounters a white pixel, for each pixel
# we encounter this way, we write the corresponding pixel in the shape array black. In doing this, we outline the
# sample within each image, which can then be used to calculate porosity.
def shape_outliner(image):
	height = image.shape[0]
	width = image.shape[1]
	shape = np.ones((height, width), dtype=int)
	cur_i = image

	# Move right to left across image moving down
	for j in range(0, height):
		width_index = width - 1  # Right hand side of image
		height_index = j  # Top of image
		if np.count_nonzero(cur_i[height_index, 0:width]) < 2:  # Check if there are any white pixels in current row
			shape[height_index, 0:width] = 0
		else:
			while cur_i[
				height_index, width_index] == 0 and width_index > width_inc:  # Loop until white pixel or boundary found
				width_index -= width_inc
			shape[height_index,
			width_index + 1:width] = 0  # Color the pixels we just iterated through black in shape array

	# Move left to right moving down
	for j in range(0, height):
		width_index = 0  # Left hand side of image
		height_index = j  # Top of image
		if np.count_nonzero(cur_i[height_index, 0:width]) < 2:  # Check if there are any white pixels in current row
			shape[height_index, 0:width] = 0
		else:
			while cur_i[
				height_index, width_index] == 0 and width_index < width - width_inc:  # Loop until white pixel or boundary found
				width_index += width_inc
			shape[height_index, 0:width_index] = 0  # Color the pixels we just iterated through black in shape array

	return shape


# Returns the porosity given a count of bright pixels vs amount of total pixels
def por_calc(bright_pixels, total_pixels):
	return (1 - (bright_pixels / float(total_pixels))) * 100


# Counts the volume for a CT scan (Multiply the surface area by voxel size) Returns volume in cm^3 (I think)
# NOTE: Assumes no porosity in volume calculation, for 'real' volume, simply subtract volume*porosity from this volume
def count_volume(total_voxels):
	volume = total_voxels * voxel_size
	return volume


# Given a excel sheet, data_writer writes the porosity previously computed to the excel sheet with a clean format
def data_writer(ws, porosity, volume):
	i = 1  # We loop to find empty cells before outputting the results
	while ws.cell(row=i, column=1).value is not None and ws.cell(row=i + 2, column=1).value is not None:
		i += 4
	ws.cell(row=i, column=1).value = "Trial %d" % (i // 3)
	ws.cell(row=i + 1, column=1).value = "Porosity:"
	ws.cell(row=i + 1, column=2).value = porosity
	ws.cell(row=i + 2, column=1).value = "Volume:"
	ws.cell(row=i + 2, column=2).value = volume


# Displays the progress of the program (percent wise) to the terminal. Inspired by:
# https://stackoverflow.com/questions/43515165/pycharm-python-console-printing-on-the-same-line-not-working-as-intended
def progress_tracker(completion, total, operation):
	sys.stdout.write("\r{0}".format(operation + str("%.2f" % (completion / total * 100)) + "% Complete"))
	sys.stdout.flush()
	if completion == total and operation == 'Porosity estimation ':
		print("\nDONE!")


# Creates the directory where I will be outputting the images, automatically creates a new directory for each new run
# of the app, all stored within the PyPore master folder, hopefully to promote organization.
def output_images():
	# Try to create the main directory containing the sub directories
	try:
		os.makedirs(r'%s/PyPore_Images/' % file_path)
	except FileExistsError:
		pass

	folder_count = 1
	made_folder = False
	output_directory = '%s/PyPore_Images/' % file_path

	# Iterate until I can make a sub directory for my images to be saved too
	while not made_folder:
		try:
			os.makedirs(r'%s/PyPore_Images_%d/' % (output_directory, folder_count))
			made_folder = True
		except FileExistsError:
			folder_count += 1

	return '%sPyPore_Images_%d' % (output_directory, folder_count)  # Location where images will be saved too


##########################################Creating loading screen#######################################################


# This function uses threading to display the loading bar and run the backend at the same time
def loading_backend():
	global loading_text, root

	# Creates the loading screen and keeps it running until the backend is finished
	root = Tk()
	loading_text = StringVar(root)  # Make the loading bar text a malleable variable
	progress_window(root)

	Thread()  # Thread class runs the backend

	while still_loading:
		loading_text.set(current_operation)
		root.update_idletasks()
		root.update()
	root.destroy()


# Modified from https://stackoverflow.com/questions/459083/how-do-you-run-your-own-code-alongside-tkinters-event-loop/538559#538559
# Using a thread class to run the backend.
class Thread(threading.Thread):

	def __init__(self):
		threading.Thread.__init__(self)
		self.start()

	def run(self):
		global still_loading

		main_flow()
		still_loading = False  # When main flow is done executing we kill the loading screen using this global variable


# A Tkinter frame that displays a loading bar to let the user know the system is doing computation.
class progress_window(Frame):

	def __init__(self, master=None):
		Frame.__init__(self, master)
		self.master = master
		self.grid(sticky=(E + W + N + S))

		master.title("PyPore")
		master.iconbitmap('PyPore.ico')
		master.resizable(width=False, height=False)
		master.columnconfigure(0, weight=1)
		master.geometry("350x100")

		self.master.progress = Progressbar(self, orient=HORIZONTAL, length=345, mode='determinate')
		self.master.progress_label = Label(self, textvariable=loading_text)
		self.master.progress_label.grid(row=0, column=0, padx=100, pady=(20, 0))
		self.master.progress.grid(row=1, column=0, padx=3)
		self.master.progress.start()



