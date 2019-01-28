# A porosity calculator designed to work with essentially any CT scan, assuming that the sample is relatively closed.
# Works using a shrinkwrap ROI approach to delineate the boundary of our sample within the image. Given a boundary,
# computing porosity simply involves taking a ratio of dark/light pixels. Does not produce accurate results if: images
# are in any color other than black and white, have lots of noisy pixels, sample has large gaps in its boundary. This
# version includes the excel file writer, better documentation and a temporary fix for the zero division problem.

import glob
import cv2
import numpy as np
import matplotlib.pyplot as plt
import random
import sys
import time
import openpyxl
from openpyxl import load_workbook

file_location = None
loop_inc = 1  # Represents the step for the main loop, skips loop_inc - 1 images (Increase to speed up runtime)
width_inc = 1  # Represents the inc/decrement step in shape outliner (Increase to speed up runtime)
save_count = 0  # Keeps track of how many comparison images are saved (for naming purposes)


# Main control for the program. Reads in images and iterates them through shape_outliner and then por_calc
def main():
	wb_ws_save = excel_handler()  # Setup excel sheet
	wb = wb_ws_save[0]  # The current workbook
	ws = wb_ws_save[1]  # The current sheet within the workbook
	save_as = wb_ws_save[2]  # The title of the current workbook

	images = file_reader()  # Read in images
	porosities = []

	start = time.time()

	# Iterate through the images and count the porosity
	for i in range(0, len(images), loop_inc):
		progress_tracker(i + 1, len(images))  # Shows user the programs progress
		total_img_pixels = np.count_nonzero(images[i])
		if total_img_pixels != 0:  # Exclude all-black images
			shape = shape_outliner(images[i])
			total_shape_pixels = np.count_nonzero(shape)
			if total_shape_pixels != 0:  # Because some all black images are "sneaking through", do a second check!
				porosities.append(por_calc(total_img_pixels, total_shape_pixels))

	if len(porosities) != 0:
		porosity = np.average(np.array(porosities))
	else:
		porosity = 0

	data_writer(ws, porosity)
	wb.save(save_as)

	print("Porosity: " + str(porosity))

	end = time.time()
	print("Time: " + str(end - start))


# Reads in image files specified by the users
def file_reader():
	global file_location

	while True:
		file_location = input("Please specify the file path to where your images are stored: ")
		file_type = input("Please specify the file type. I.e bmp: ")
		files = (glob.glob(file_location + "/*." + file_type))  # Read all files in a folder using glob
		images = []
		for img_file in files:
			images.append(cv2.imread(img_file, 0))  # Turn the files into an array of images for easier access
		if len(images) == 0:
			print("The folder location or file type you selected were invalid, please try again")
			continue
		break

	return images


# Loads in a pre existing excel file or creates a new one for writing based on user input
def excel_handler():
	valid_file = False

	while not valid_file:
		old_excel = input("Would you like to use an existing excel file Y OR N: ")
		if old_excel == "Y" or old_excel == "y":
			excel_file = input("Please specify the file name with proper suffix i.e sheet1.xlsx: ")
			try:
				wb = load_workbook(excel_file)
				ws = wb.active
				save_as = excel_file
				valid_file = True
			except IOError:
				print("Your specified file does not exist, please try again")

		elif old_excel == "N" or old_excel == "n":
			valid_file = True
			wb = openpyxl.Workbook()
			save_as = input("What would you like to save the excel book as: ")
			save_as = save_as + ".xlsx"
			for chars in save_as:
				if chars in [':', '*', '?', '"', '<', '>', '|'] or ord(chars) == 92 or ord(chars) == 47:
					print("This is not a valid file name, it contains a :,*,?,<,>,|,/ etc. ")
					valid_file = False
			ws = wb.active

		else:
			print("Sorry, I didn't understand that.")
			continue

	return wb, ws, save_as


# Given a excel sheet, data_writer writes the porosity previously computed to the excel sheet with a clean format
def data_writer(ws, porosity):
	i = 1  # We loop to find empty cells before outputting the results
	while ws.cell(row=i, column=1).value is not None and ws.cell(row=i + 1, column=1).value is not None:
		i += 3
	ws.cell(row=i, column=1).value = "Trial %d" % (i // 3)
	ws.cell(row=i + 1, column=1).value = "Porosity:"
	ws.cell(row=i + 1, column=2).value = porosity


# Randomly saves an image based on a tunable probability "odds" within functions
def random_saver():
	odds = 10  # Think of this as probability to save = 1/odds
	x = random.randrange(odds)
	return x == odds - 1


# Saves a plot showing a image and its corresponding shape outline, allows user to ensure proper operation of code
def image_saver(image, shape):
	global save_count

	plt.clf()
	plt.subplot(2, 1, 1)
	plt.title("Original Image")
	plt.imshow(image, cmap="gray")
	plt.subplot(2, 1, 2)
	plt.title("Shape outlined Image")
	plt.imshow(shape, cmap="gray")
	plt.savefig(file_location + "\Comparison_Number_" + str(save_count) + ".png")
	save_count += 1


# Displays the progress of the program (percent wise) to the terminal. Inspired by:
# https://stackoverflow.com/questions/43515165/pycharm-python-console-printing-on-the-same-line-not-working-as-intended
def progress_tracker(completion, total):
	sys.stdout.write("\r{0}".format(str("%.2f" % (completion / total * 100)) + "% Complete"))
	sys.stdout.flush()
	if completion == total:
		print("\nDONE!")


# Mimics a ROI shrink wrap procedure. Traces across original image until it encounters a white pixel, for each pixel
# we encounter this way, we write the corresponding pixel in the shape array black. In doing this, we outline the
# sample within each image, which can then be used to calculate porosity.
def shape_outliner(image):

	height = image.shape[0]
	width = image.shape[1]
	shape = np.ones((height, width), dtype=int)
	cur_i = image

	# Move right to left across image moving down
	for j in range(0, height):
		width_index = width - 1  # Right hand side of image
		height_index = j  # Top of image
		while cur_i[height_index, width_index] == 0 and width_index > width_inc:  # Loop until white pixel or boundary found
			width_index -= width_inc
		shape[height_index, width_index:width] = 0  # Color the pixels we just iterated through black in shape array

	# Move left to right moving down
	for j in range(0, height):
		width_index = 0  # Left hand side of image
		height_index = j  # Top of image
		while cur_i[height_index, width_index] == 0 and width_index < width - width_inc:  # Loop until white pixel or boundary found
			width_index += width_inc
		shape[height_index, 0:width_index] = 0  # Color the pixels we just iterated through black in shape array

	shape = cv2.medianBlur(shape.astype(np.float32), 5)  # Sooth images to remove border irregularities

	# Randomly choose to save a comparison image (Ensure shape outliner works properly)
	if random_saver():
		image_saver(image, shape)

	return shape


# Returns the porosity given a count of bright pixels vs amount of total pixels
def por_calc(image_bright_pixels, shape_bright_pixels):
	return (1 - (image_bright_pixels/float(shape_bright_pixels))) * 100


main()
