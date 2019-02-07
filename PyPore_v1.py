import cv2
import glob
import random
import re
import sys
import numpy as np
import openpyxl
from openpyxl import load_workbook
from skimage.morphology import remove_small_objects
from scipy.ndimage import label

file_location = None
width_inc = 1  # Represents the inc/decrement step in shape outliner (Increase to speed up runtime)


# Main control for the program. Reads in images, creates/loads in excel file, calls functions that preform image
# processing (Thresholding, Despeckeling and Cropping) and image analysis (Measure porosity and volume). Main also
# saves the results to an excel file and includes a timer for testing the applications performance.
def main():
	images = file_reader()  # Read in images

	wb_ws_save = excel_handler()  # Setup excel sheet
	wb = wb_ws_save[0]  # The current workbook
	ws = wb_ws_save[1]  # The current sheet within the workbook
	save_as = wb_ws_save[2]  # The title of the current workbook

	valid_voxel_size = False

	while not valid_voxel_size:
		voxel_size = input("What is the voxel size for your images in um: ")
		try:
			voxel_size = np.float64(voxel_size) * (10 ** -4)  # Converting um to cm
			voxel_size = voxel_size ** 3  # Cubing cm
			print(voxel_size)
			valid_voxel_size = True
		except ValueError:
			print("Please enter an integer")

	images = image_processor(images)

	results = analyze(images)  # Results computes porosity/ surface area for each slice
	porosities = results[0]
	surface_area = results[1]

	# Ensure we do not get error taking average on empty array
	if len(porosities) != 0:
		porosity = np.average(np.array(porosities))
	else:
		porosity = 0

	volume = count_volume(surface_area, voxel_size)

	data_writer(ws, porosity, volume)  # Write the porosity to the excel file
	wb.save(save_as)  # Saves the excel file


# Analyze preforms porosity and surface area measurements on binary images. First each image gets a ROI shrinkwrap
# which accurately delineates the sample boundary. The total pixels of this shrinkwrap are added a surface area holder
# which is later multiplied by voxel size to estimate object volume. Likewise a ratio between bright pixels in the
# shape outlined image vs original image is used to estimate porosity. These values are then returned to main.
def analyze(images):
	porosities = []
	surface_area = 0

	# Iterate through the images counting porosity and surface area for each slice
	for i in range(0, len(images)):
		progress_tracker(i + 1, len(images), 'Porosity estimation ')  # Shows user the programs progress
		total_img_pixels = np.count_nonzero(images[i])
		if total_img_pixels > 10:  # Exclude images that contain little to no white pixels
			shape = shape_outliner(images[i])
			total_shape_pixels = np.count_nonzero(shape)
			porosities.append(por_calc(total_img_pixels, total_shape_pixels))
			surface_area += total_shape_pixels

	return porosities, surface_area


# Image processors handles all image processing procedures which includes thresholding, despeckeling and cropping.
# The despeckle and crop functionality are essentially complete, but may need to be reworked with UI implementation
def image_processor(images):
	# Threshold all images (Currently only Otsus Threshold)
	for i in range(0, len(images)):
		progress_tracker(i + 1, len(images), 'Thresholding ')  # Shows user the programs progress
		images[i] = thresholder(images[i])

	# Despeckle all images, 3 option, speckles less than, greater than or auto despeckle
	despeckle_options = input("\nYour options for despeckeling are as follows \n1) Remove white specks less than x\n2) "
	"Remove white specks greater than x\n3) Auto-Despeckle\n4) Skip\nChoose your option: ")
	if despeckle_options == "1":
		min_area = int(input("Remove speckles of size less than: "))
	elif despeckle_options == "2":
		min_area = int(input("Remove speckles of size greater than: "))
	elif despeckle_options == "3":
		min_area = auto_despeckle_parameters(images)

	if despeckle_options != "4":
		for i in range(0, len(images)):
			progress_tracker(i + 1, len(images), 'Despeckeling ')  # Shows user the programs progress
			despeckle_img = despeckle(images[i], min_area)
			if despeckle_options == "2":  # Using XOR to simulate despeckle of objects greater than X
				images[i] = np.logical_xor(images[i], despeckle_img)
			else:
				images[i] = despeckle_img

	# Crop all images
	parameters = crop_parameters(images)  # Get the universal parameters for future cropping operations
	for i in range(0, len(images)):
		progress_tracker(i + 1, len(images), 'Cropping ')  # Shows user the programs progress
		images[i] = crop(images[i], parameters)

	return images


# Reads in image files specified by the users
def file_reader():
	global file_location

	while True:
		file_location = input("Please specify the file path to where your images are stored: ")
		file_type = input("Please specify the file type. I.e bmp: ")
		files = (glob.glob(file_location + "/*." + file_type))  # Read all files in a folder using glob
		images = []
		for img_file in files:
			images.append(cv2.imread(img_file, 0))  # Turn the files into an array of images for easier access
		if len(images) == 0:
			print("The folder location or file type you selected were invalid, please try again")
			continue
		break

	return images


# Loads in a pre existing excel file or creates a new one for writing based on user input
def excel_handler():
	valid_file = False

	while not valid_file:
		old_excel = input("Would you like to use an existing excel file Y OR N: ")
		if old_excel == "Y" or old_excel == "y":  # User has selected a pre-existing workbook
			excel_file = file_location + "/" + input("Please specify the excel file name: ") + ".xlsx"
			try:
				wb = load_workbook(excel_file)
				ws = wb.active
				save_as = excel_file
				valid_file = True
			except IOError:
				print("Your specified file does not exist, please try again")

		elif old_excel == "N" or old_excel == "n":  # User has selected to create a new workbook
			wb = openpyxl.Workbook()
			save_as = input("What would you like to save the excel book as: ")
			if re.match("^[A-Za-z0-9_+@#^&()_,.!-]*$", save_as):
				save_as = file_location + "/" + save_as + ".xlsx"
				ws = wb.active
				valid_file = True
			else:
				print("This is not a valid file name, it contains a :,*,?,[,],\,/... please try again ")

		else:
			print("Sorry, I didn't understand that.")
			continue

	return wb, ws, save_as


# Given a excel sheet, data_writer writes the porosity previously computed to the excel sheet with a clean format
def data_writer(ws, porosity, volume):
	i = 1  # We loop to find empty cells before outputting the results
	while ws.cell(row=i, column=1).value is not None and ws.cell(row=i + 2, column=1).value is not None:
		i += 4
	ws.cell(row=i, column=1).value = "Trial %d" % (i // 3)
	ws.cell(row=i + 1, column=1).value = "Porosity:"
	ws.cell(row=i + 1, column=2).value = porosity
	ws.cell(row=i + 2, column=1).value = "Volume:"
	ws.cell(row=i + 2, column=2).value = volume


# Displays the progress of the program (percent wise) to the terminal. Inspired by:
# https://stackoverflow.com/questions/43515165/pycharm-python-console-printing-on-the-same-line-not-working-as-intended
def progress_tracker(completion, total, operation):
	sys.stdout.write("\r{0}".format(operation + str("%.2f" % (completion / total * 100)) + "% Complete"))
	sys.stdout.flush()
	if completion == total and operation == 'Porosity estimation ':
		print("\nDONE!")


# Given a set of images, crop parameters estimates an appropriate crop boundary. To do this, we first consider x images
# with the highest bright pixel count. We assume there is a correlation between # of bright pixels and the area of the
# sample in the image. Given these top images, we then test how far we can crop each image in each of the compass
# directions. We then take the smallest values from each direction and then apply a scaling factor to these values.
# By doing this computation, we are careful to not cut into our data when doing our cropping.
def crop_parameters(images):
	height = images[0].shape[0]
	width = images[0].shape[1]

	images_considered = 10
	if len(images) < images_considered:  # Ensures we do not try and consider more images then exist
		images_considered = len(images)

	top_images = np.zeros([images_considered, 2], np.int32)
	parameters = np.zeros([images_considered, 4], np.int32)
	final_parameters = np.zeros([1, 4], np.int32)

	# Approximates the images with the largest respective area (assume most white pixels = most area)
	for i in range(0, len(images)):
		if np.count_nonzero(images[i]) > min(top_images[:, 1]):
			top_images[np.argmin(top_images[:, 1])] = [i, np.count_nonzero(images[i])]

	# Generate the number of pixels required to reach the sample in all 4 directions which guides where to crop
	for i in range(0, images_considered):
		parameters[i, 0] = top_down_shutter_close(images[top_images[i, 0]], 0, width, 1)  # Top Down
		parameters[i, 1] = top_down_shutter_close(images[top_images[i, 0]], height - 1, width, -1)  # Bottom Up
		parameters[i, 2] = left_right_shutter_close(images[top_images[i, 0]], 0, height, 1)  # Left to Right
		parameters[i, 3] = left_right_shutter_close(images[top_images[i, 0]], width - 1, height, -1)  # Right to Left

	# Take the smallest value for each of the 4 directions (minimum distance before hitting data)
	for i in range(0, 4):
		final_parameters[0, i] = min(parameters[:, i])

	return final_parameters[0]


# Used to probe how far a 2D plane can go in the image before reaching the dataset, works from the top and bottom edges
def top_down_shutter_close(image, height_index, width, increment):
	counter = 0

	# The loop ensure we do not stop at a faulty value caused by image noise, must be more than 10 non zero pixels in plane
	while np.count_nonzero(image[height_index, 0:width]) < 10 and height_index < image.shape[0]:
		height_index += increment
		counter += 1

	return counter


# Used to probe how far a 2D plane can go in the image before reaching the dataset, works from the left and right edges
def left_right_shutter_close(image, width_index, height, increment):
	counter = 0

	# The loop ensure we do not stop at a faulty value caused by image noise, must be more than 10 non zero pixels in plane
	while np.count_nonzero(image[0:height, width_index]) < 10 and width_index < image.shape[1]:
		width_index += increment
		counter += 1

	return counter


# Given an image and a list of parameters, crops the image to a smaller dimension for faster future processing
def crop(img, dimensions):
	scale = 1
	crop_img = img[int(dimensions[0] // scale):img.shape[0] - int(dimensions[1] // scale),
	int(dimensions[2] // scale):img.shape[1] - int(dimensions[3] // scale)]

	return crop_img


# Mimics a ROI shrink wrap procedure. Traces across original image until it encounters a white pixel, for each pixel
# we encounter this way, we write the corresponding pixel in the shape array black. In doing this, we outline the
# sample within each image, which can then be used to calculate porosity.
def shape_outliner(image):
	height = image.shape[0]
	width = image.shape[1]
	shape = np.ones((height, width), dtype=int)
	cur_i = image

	# Move right to left across image moving down
	for j in range(0, height):
		width_index = width - 1  # Right hand side of image
		height_index = j  # Top of image
		if np.count_nonzero(cur_i[height_index, 0:width]) < 2:  # Check if there are any white pixels in current row
			shape[height_index, 0:width] = 0
		else:
			while cur_i[
				height_index, width_index] == 0 and width_index > width_inc:  # Loop until white pixel or boundary found
				width_index -= width_inc
			shape[height_index,
			width_index + 1:width] = 0  # Color the pixels we just iterated through black in shape array

	# Move left to right moving down
	for j in range(0, height):
		width_index = 0  # Left hand side of image
		height_index = j  # Top of image
		if np.count_nonzero(cur_i[height_index, 0:width]) < 2:  # Check if there are any white pixels in current row
			shape[height_index, 0:width] = 0
		else:
			while cur_i[
				height_index, width_index] == 0 and width_index < width - width_inc:  # Loop until white pixel or boundary found
				width_index += width_inc
			shape[height_index, 0:width_index] = 0  # Color the pixels we just iterated through black in shape array

	# shape = cv2.medianBlur(shape.astype(np.float32), 5)  # Sooth images to remove border irregularities
	# In really low porosity samples can lead to negative porosity values, uncomment with caution.

	return shape


# Returns the porosity given a count of bright pixels vs amount of total pixels
def por_calc(bright_pixels, total_pixels):
	return (1 - (bright_pixels / float(total_pixels))) * 100


# Apply Otsus threshold to all images
def thresholder(image):
	null, thres_img = cv2.threshold(image, 125, 1, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
	return thres_img


# Auto despeckle is designed to work with single object, low porosity samples (i.e salt scans). It works by
# iteratively increasing the min object despeckeled until only one object remains. We repeat this process on random
# images and take the average value after x trials.
def auto_despeckle_parameters(images):
	images_considered = 10
	min_area_array = np.zeros([images_considered])

	for i in range(0, images_considered):
		progress_tracker(i + 1, images_considered, 'Starting Auto Crop ')
		random_img = images[random.randrange(len(images))]  # Select a random image from the stack
		bool_arr = np.array(random_img, bool)  # Convert it into a boolean array (needed for small objects method)

		min_obj_size = 0
		num_features = -1

		# Loop increasing min object size until only one object is left in the image
		while num_features != 1 and min_obj_size < 100:
			print(num_features)
			min_obj_size += 10
			filtered_img = remove_small_objects(bool_arr, min_obj_size)
			null, num_features = label(filtered_img)

		min_area_array[i] = min_obj_size

	return int(np.average(min_area_array))


# Despeckle images using a remove small objects function, size of objects is either user determined or found
# automatically via random experimentation in auto crop parameters
def despeckle(image, min_area):
	bool_arr = np.array(image, bool)
	despeckeled_img = remove_small_objects(bool_arr, min_area)
	return np.array(despeckeled_img, dtype=int)


# Counts the volume for a CT scan (Multiply the surface area by voxel size) Returns volume in cm^3 (I think)
# NOTE: Assumes no porosity in volume calculation, for 'real' volume, simply subtract volume*porosity from this volume
def count_volume(total_voxels, voxel_size):
	volume = total_voxels * voxel_size
	return volume


main()
