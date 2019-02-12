import cv2
import glob
import random
import re
import sys
import numpy as np
import openpyxl
from openpyxl import load_workbook
from skimage.morphology import remove_small_objects
from scipy.ndimage import label

img_files_location = None
test_image = None
globbed_image = None
workbook = None
save_excel_file_as = None
threshold_type = None
despeckle_type = None
width_inc = 1  # Represents the inc/decrement step in shape outliner (Increase to speed up runtime)


def main_flow():
	pass
	# Read in all images
	images = []
	for img_file in globbed_image:
		images.append(cv2.imread(img_file, 0))  # Turn the files into an array of images for easier access

	# Preform Thresholding

	# Preform Despeckeling
	# Preform AutoCropping
	# Call Analyzer


# TODO NAYBE MULTI THREAD, IF NOT PROBABLY BEST TO LOAD ALL IMAGES ANYWAYS
def file_reader(file_path, file_type):
	global img_files_location, test_image, globbed_image
	file_location = file_path + "/*." + file_type
	globbed_image = (glob.glob(file_location))  # Read all files in a folder using glob

	try:
		test_image = (cv2.imread(globbed_image[len(globbed_image)//2], 0))  # Try to read a file
		img_files_location = file_path
		return True
	except IndexError:
		return False


def excel_reader(file_location, new_sheet):
	global workbook
	global save_excel_file_as

	full_file_path = img_files_location + "/" + file_location + ".xlsx"
	print(full_file_path)

	if new_sheet:
		try:
			workbook = load_workbook(full_file_path)
			save_excel_file_as = file_location
			# ws = wb.active TODO DO THIS IN EXCEL FILE READER
			return 0
		except IOError:
			print("Your specified file does not exist, please try again")
			return 1

	else:
		try:
			load_workbook(full_file_path)
			print("Attempting to overwrite workbook")
			return 2
		except IOError:
			if re.match("^[A-Za-z0-9_+@#^&(),.!-]*$", file_location):
				workbook = openpyxl.Workbook()
				save_excel_file_as = file_location
				return 0
			else:
				return 3


def threshold_choice(user_choice):
	global threshold_type
	if threshold_type is None:
		threshold_type = user_choice


def threshold_selector(user_choice):
	if user_choice == 1:
		return otsu_threshold()
	elif user_choice == 2:
		return phanstalker_threshold()
	elif user_choice == 3:
		return global_threshold()


# TODO IMPLEMENT THE OTHER THRESHOLDING TECHNIQUES
def otsu_threshold(image=None):
	if image is None:
		image = test_image

	null, thres_img = cv2.threshold(image, 125, 1, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
	return image


def phanstalker_threshold(image=None):
	if image is None:
		image = test_image

	null, thres_img = cv2.threshold(image, 125, 1, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
	return thres_img


def global_threshold(image=None):
	if image is None:
		image = test_image

	null, thres_img = cv2.threshold(image, 125, 1, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
	return thres_img


def despeckle_choice(user_choice, area=None):
	global despeckle_type
	if despeckle_type is None:
		despeckle_type = user_choice, area


def despeckle_selector(user_choice, area):
	if user_choice == 1:
		return despeckle(test_image, area)
	elif user_choice == 2:
		return np.logical_xor(test_image, despeckle(test_image, area))


# Despeckle images using a remove small objects function, size of objects is either user determined or found
# automatically via random experimentation in auto crop parameters
def despeckle(image, min_area):
	bool_arr = np.array(image, bool)
	despeckeled_img = remove_small_objects(bool_arr, min_area)
	return np.array(despeckeled_img, dtype=int)


# Given a set of images, crop parameters estimates an appropriate crop boundary. To do this, we first consider x images
# with the highest bright pixel count. We assume there is a correlation between # of bright pixels and the area of the
# sample in the image. Given these top images, we then test how far we can crop each image in each of the compass
# directions. We then take the smallest values from each direction and then apply a scaling factor to these values.
# By doing this computation, we are careful to not cut into our data when doing our cropping.
def crop_parameters(images):
	height = images[0].shape[0]
	width = images[0].shape[1]

	images_considered = 10
	if len(images) < images_considered:  # Ensures we do not try and consider more images then exist
		images_considered = len(images)

	top_images = np.zeros([images_considered, 2], np.int32)
	parameters = np.zeros([images_considered, 4], np.int32)
	final_parameters = np.zeros([1, 4], np.int32)

	# Approximates the images with the largest respective area (assume most white pixels = most area)
	for i in range(0, len(images)):
		if np.count_nonzero(images[i]) > min(top_images[:, 1]):
			top_images[np.argmin(top_images[:, 1])] = [i, np.count_nonzero(images[i])]

	# Generate the number of pixels required to reach the sample in all 4 directions which guides where to crop
	for i in range(0, images_considered):
		parameters[i, 0] = top_down_shutter_close(images[top_images[i, 0]], 0, width, 1)  # Top Down
		parameters[i, 1] = top_down_shutter_close(images[top_images[i, 0]], height - 1, width, -1)  # Bottom Up
		parameters[i, 2] = left_right_shutter_close(images[top_images[i, 0]], 0, height, 1)  # Left to Right
		parameters[i, 3] = left_right_shutter_close(images[top_images[i, 0]], width - 1, height, -1)  # Right to Left

	# Take the smallest value for each of the 4 directions (minimum distance before hitting data)
	for i in range(0, 4):
		final_parameters[0, i] = min(parameters[:, i])

	return final_parameters[0]


# Used to probe how far a 2D plane can go in the image before reaching the dataset, works from the top and bottom edges
def top_down_shutter_close(image, height_index, width, increment):
	counter = 0

	# The loop ensure we do not stop at a faulty value caused by image noise, must be more than 10 non zero pixels in plane
	while np.count_nonzero(image[height_index, 0:width]) < 10 and height_index < image.shape[0]:
		height_index += increment
		counter += 1

	return counter


# Used to probe how far a 2D plane can go in the image before reaching the dataset, works from the left and right edges
def left_right_shutter_close(image, width_index, height, increment):
	counter = 0

	# The loop ensure we do not stop at a faulty value caused by image noise, must be more than 10 non zero pixels in plane
	while np.count_nonzero(image[0:height, width_index]) < 10 and width_index < image.shape[1]:
		width_index += increment
		counter += 1

	return counter


# Given an image and a list of parameters, crops the image to a smaller dimension for faster future processing
def crop(img, dimensions):
	scale = 1
	crop_img = img[int(dimensions[0] // scale):img.shape[0] - int(dimensions[1] // scale),
	int(dimensions[2] // scale):img.shape[1] - int(dimensions[3] // scale)]

	return crop_img


# Mimics a ROI shrink wrap procedure. Traces across original image until it encounters a white pixel, for each pixel
# we encounter this way, we write the corresponding pixel in the shape array black. In doing this, we outline the
# sample within each image, which can then be used to calculate porosity.
def shape_outliner(image):
	height = image.shape[0]
	width = image.shape[1]
	shape = np.ones((height, width), dtype=int)
	cur_i = image

	# Move right to left across image moving down
	for j in range(0, height):
		width_index = width - 1  # Right hand side of image
		height_index = j  # Top of image
		if np.count_nonzero(cur_i[height_index, 0:width]) < 2:  # Check if there are any white pixels in current row
			shape[height_index, 0:width] = 0
		else:
			while cur_i[
				height_index, width_index] == 0 and width_index > width_inc:  # Loop until white pixel or boundary found
				width_index -= width_inc
			shape[height_index,
			width_index + 1:width] = 0  # Color the pixels we just iterated through black in shape array

	# Move left to right moving down
	for j in range(0, height):
		width_index = 0  # Left hand side of image
		height_index = j  # Top of image
		if np.count_nonzero(cur_i[height_index, 0:width]) < 2:  # Check if there are any white pixels in current row
			shape[height_index, 0:width] = 0
		else:
			while cur_i[
				height_index, width_index] == 0 and width_index < width - width_inc:  # Loop until white pixel or boundary found
				width_index += width_inc
			shape[height_index, 0:width_index] = 0  # Color the pixels we just iterated through black in shape array

	return shape


# Returns the porosity given a count of bright pixels vs amount of total pixels
def por_calc(bright_pixels, total_pixels):
	return (1 - (bright_pixels / float(total_pixels))) * 100


# Analyze preforms porosity and surface area measurements on binary images. First each image gets a ROI shrinkwrap
# which accurately delineates the sample boundary. The total pixels of this shrinkwrap are added a surface area holder
# which is later multiplied by voxel size to estimate object volume. Likewise a ratio between bright pixels in the
# shape outlined image vs original image is used to estimate porosity. These values are then returned to main.
def analyze(images):
	porosities = []
	surface_area = 0

	# Iterate through the images counting porosity and surface area for each slice
	for i in range(0, len(images)):
		# progress_tracker(i + 1, len(images), 'Porosity estimation ')  # Shows user the programs progress
		total_img_pixels = np.count_nonzero(images[i])
		if total_img_pixels > 10:  # Exclude images that contain little to no white pixels
			shape = shape_outliner(images[i])
			total_shape_pixels = np.count_nonzero(shape)
			porosities.append(por_calc(total_img_pixels, total_shape_pixels))
			surface_area += total_shape_pixels

	return porosities, surface_area
